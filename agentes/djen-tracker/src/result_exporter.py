"""
Result Exporter - Exportação Multi-Formato de Resultados

Exporta resultados de filtro OAB em múltiplos formatos:
- JSON: Estruturado para processamento automatizado
- Markdown: Formatado para leitura humana
- TXT: Simples para parsing posterior
- Excel: Tabela com formatação condicional (opcional)
- HTML: Relatório visual (opcional)

Author: Claude Code (Development Agent)
Version: 1.0.0
"""

import json
import logging
from pathlib import Path
from typing import List, Optional, Dict, Tuple
from datetime import datetime
from collections import defaultdict


logger = logging.getLogger(__name__)


class ResultExporter:
    """
    Exportador de resultados em múltiplos formatos.

    Attributes:
        group_by_tribunal: Agrupar resultados por tribunal
        group_by_oab: Agrupar resultados por OAB
    """

    def __init__(
        self,
        group_by_tribunal: bool = True,
        group_by_oab: bool = False
    ):
        """
        Inicializa ResultExporter.

        Args:
            group_by_tribunal: Agrupar por tribunal
            group_by_oab: Agrupar por OAB
        """
        self.group_by_tribunal = group_by_tribunal
        self.group_by_oab = group_by_oab

    def export_json(
        self,
        matches: List,  # List[PublicacaoMatch]
        output_path: Path,
        indent: int = 2
    ) -> None:
        """
        Exporta para JSON estruturado.

        Args:
            matches: Lista de PublicacaoMatch
            output_path: Path para arquivo de saída
            indent: Indentação do JSON
        """
        logger.info(f"Exportando {len(matches)} matches para JSON...")

        data = {
            'metadata': {
                'total_matches': len(matches),
                'exported_at': datetime.now().isoformat(),
                'format_version': '1.0'
            },
            'matches': [match.to_dict() for match in matches]
        }

        # Agrupar se solicitado
        if self.group_by_tribunal:
            grouped = self._group_by_tribunal(matches)
            data['grouped_by_tribunal'] = {
                tribunal: [m.to_dict() for m in matches_list]
                for tribunal, matches_list in grouped.items()
            }

        if self.group_by_oab:
            grouped = self._group_by_oab(matches)
            data['grouped_by_oab'] = {
                f"{num}/{uf}": [m.to_dict() for m in matches_list]
                for (num, uf), matches_list in grouped.items()
            }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)

        logger.info(f"JSON exportado: {output_path}")

    def export_markdown(
        self,
        matches: List,  # List[PublicacaoMatch]
        output_path: Path,
        include_toc: bool = True
    ) -> None:
        """
        Exporta para Markdown formatado.

        Args:
            matches: Lista de PublicacaoMatch
            output_path: Path para arquivo de saída
            include_toc: Incluir índice
        """
        logger.info(f"Exportando {len(matches)} matches para Markdown...")

        lines = []

        # Cabeçalho
        lines.append("# Publicações OAB - Relatório de Filtro\n")
        lines.append(f"**Gerado em:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lines.append(f"**Total de publicações:** {len(matches)}\n")
        lines.append("---\n")

        # Estatísticas
        lines.append("## Estatísticas\n")
        stats = self._calculate_stats(matches)
        lines.append(f"- **Tribunais:** {stats['total_tribunais']}")
        lines.append(f"- **OABs únicas:** {stats['total_oabs']}")
        lines.append(f"- **Score médio:** {stats['score_medio']:.2f}")
        lines.append(f"- **Publicações com revisão manual:** {stats['requer_revisao']}")
        lines.append("\n")

        # Distribuição por tribunal
        lines.append("### Distribuição por Tribunal\n")
        for tribunal, count in stats['por_tribunal'].items():
            lines.append(f"- **{tribunal}:** {count} publicações")
        lines.append("\n")

        # Índice (se solicitado)
        if include_toc:
            lines.append("## Índice\n")
            for i, match in enumerate(matches, 1):
                lines.append(
                    f"{i}. [{match.tribunal} - {match.data_publicacao}] "
                    f"OAB {match.oab_numero}/{match.oab_uf} "
                    f"(Score: {match.score_relevancia:.2f})"
                )
            lines.append("\n---\n")

        # Publicações detalhadas
        lines.append("## Publicações Detalhadas\n")

        for i, match in enumerate(matches, 1):
            lines.append(f"### {i}. Publicação {i}/{len(matches)}\n")

            # Informações principais
            lines.append(f"**Tribunal:** {match.tribunal}")
            lines.append(f"**Data:** {match.data_publicacao}")
            lines.append(f"**OAB:** {match.oab_numero}/{match.oab_uf}")
            lines.append(f"**Score de Relevância:** {match.score_relevancia:.2f}")

            # Tipo de ato
            if match.tipo_ato:
                lines.append(f"**Tipo de Ato:** {match.tipo_ato}")

            # Menções
            lines.append(f"**Total de Menções:** {match.total_mencoes}")

            # Arquivo
            lines.append(f"**Arquivo:** `{Path(match.arquivo_pdf).name}`")

            lines.append("\n**Contexto:**")
            lines.append(f"> {match.texto_contexto}")

            # Scores detalhados
            lines.append("\n**Scores Detalhados:**")
            lines.append(f"- Contexto: {match.score_contexto:.2f}")
            lines.append(f"- Densidade: {match.score_densidade:.2f}")
            lines.append(f"- Posição: {match.score_posicao:.2f}")

            # Flags
            if match.requer_revisao_manual:
                lines.append("\n⚠️ **Requer Revisão Manual**")

            lines.append("\n---\n")

        # Escrever arquivo
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        logger.info(f"Markdown exportado: {output_path}")

    def export_txt(
        self,
        matches: List,  # List[PublicacaoMatch]
        output_path: Path,
        simple_format: bool = False
    ) -> None:
        """
        Exporta para TXT simples.

        Args:
            matches: Lista de PublicacaoMatch
            output_path: Path para arquivo de saída
            simple_format: Formato simplificado (apenas essencial)
        """
        logger.info(f"Exportando {len(matches)} matches para TXT...")

        lines = []

        if simple_format:
            # Formato simples: uma linha por publicação
            lines.append(
                "# Formato: TRIBUNAL | DATA | OAB_NUMERO | UF | SCORE | TIPO_ATO | ARQUIVO\n"
            )

            for match in matches:
                lines.append(
                    f"{match.tribunal} | "
                    f"{match.data_publicacao} | "
                    f"{match.oab_numero} | "
                    f"{match.oab_uf} | "
                    f"{match.score_relevancia:.2f} | "
                    f"{match.tipo_ato or 'N/A'} | "
                    f"{Path(match.arquivo_pdf).name}"
                )

        else:
            # Formato completo
            lines.append("=" * 80)
            lines.append("PUBLICAÇÕES OAB - RELATÓRIO DE FILTRO")
            lines.append("=" * 80)
            lines.append(f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            lines.append(f"Total de publicações: {len(matches)}")
            lines.append("=" * 80)
            lines.append("")

            for i, match in enumerate(matches, 1):
                lines.append(f"[{i}/{len(matches)}] PUBLICAÇÃO")
                lines.append("-" * 80)
                lines.append(f"Tribunal: {match.tribunal}")
                lines.append(f"Data: {match.data_publicacao}")
                lines.append(f"OAB: {match.oab_numero}/{match.oab_uf}")
                lines.append(f"Score: {match.score_relevancia:.2f}")

                if match.tipo_ato:
                    lines.append(f"Tipo: {match.tipo_ato}")

                lines.append(f"Menções: {match.total_mencoes}")
                lines.append(f"Arquivo: {Path(match.arquivo_pdf).name}")
                lines.append("")
                lines.append("CONTEXTO:")
                lines.append(match.texto_contexto)

                if match.requer_revisao_manual:
                    lines.append("")
                    lines.append("⚠️  REQUER REVISÃO MANUAL")

                lines.append("")
                lines.append("=" * 80)
                lines.append("")

        # Escrever arquivo
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        logger.info(f"TXT exportado: {output_path}")

    def export_excel(
        self,
        matches: List,  # List[PublicacaoMatch]
        output_path: Path,
        include_context: bool = True
    ) -> None:
        """
        Exporta para Excel com formatação condicional.

        Requer: openpyxl

        Args:
            matches: Lista de PublicacaoMatch
            output_path: Path para arquivo de saída
            include_context: Incluir coluna de contexto
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error(
                "openpyxl não instalado. "
                "Execute: pip install openpyxl"
            )
            raise

        logger.info(f"Exportando {len(matches)} matches para Excel...")

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Publicações OAB"

        # Cabeçalhos
        headers = [
            "Nº",
            "Tribunal",
            "Data",
            "OAB Número",
            "UF",
            "Score",
            "Tipo Ato",
            "Menções",
            "Arquivo"
        ]

        if include_context:
            headers.append("Contexto")

        # Escrever cabeçalhos com estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Escrever dados
        for i, match in enumerate(matches, 1):
            row = i + 1

            data = [
                i,
                match.tribunal,
                match.data_publicacao,
                match.oab_numero,
                match.oab_uf,
                match.score_relevancia,
                match.tipo_ato or "N/A",
                match.total_mencoes,
                Path(match.arquivo_pdf).name
            ]

            if include_context:
                data.append(match.texto_contexto)

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)

                # Formatação condicional por score
                if col == 6:  # Coluna de score
                    score = match.score_relevancia
                    if score >= 0.8:
                        cell.fill = PatternFill(
                            start_color="C6EFCE",
                            end_color="C6EFCE",
                            fill_type="solid"
                        )
                    elif score >= 0.6:
                        cell.fill = PatternFill(
                            start_color="FFEB9C",
                            end_color="FFEB9C",
                            fill_type="solid"
                        )
                    else:
                        cell.fill = PatternFill(
                            start_color="FFC7CE",
                            end_color="FFC7CE",
                            fill_type="solid"
                        )

        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        if include_context:
            # Contexto mais largo
            ws.column_dimensions[get_column_letter(len(headers))].width = 50

        # Salvar
        wb.save(output_path)
        logger.info(f"Excel exportado: {output_path}")

    def export_html(
        self,
        matches: List,  # List[PublicacaoMatch]
        output_path: Path,
        include_stats: bool = True
    ) -> None:
        """
        Exporta para HTML com visualização interativa.

        Args:
            matches: Lista de PublicacaoMatch
            output_path: Path para arquivo de saída
            include_stats: Incluir estatísticas
        """
        logger.info(f"Exportando {len(matches)} matches para HTML...")

        html = []

        # Cabeçalho HTML
        html.append("<!DOCTYPE html>")
        html.append("<html lang='pt-BR'>")
        html.append("<head>")
        html.append("  <meta charset='UTF-8'>")
        html.append("  <title>Publicações OAB - Relatório</title>")
        html.append("  <style>")
        html.append(self._get_html_css())
        html.append("  </style>")
        html.append("</head>")
        html.append("<body>")

        # Título
        html.append("  <h1>Publicações OAB - Relatório de Filtro</h1>")
        html.append(
            f"  <p class='metadata'>Gerado em: "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>"
        )
        html.append(f"  <p class='metadata'>Total: {len(matches)} publicações</p>")

        # Estatísticas
        if include_stats:
            stats = self._calculate_stats(matches)
            html.append("  <div class='stats'>")
            html.append("    <h2>Estatísticas</h2>")
            html.append(f"    <p>Tribunais: {stats['total_tribunais']}</p>")
            html.append(f"    <p>OABs únicas: {stats['total_oabs']}</p>")
            html.append(f"    <p>Score médio: {stats['score_medio']:.2f}</p>")
            html.append("  </div>")

        # Tabela de publicações
        html.append("  <table>")
        html.append("    <thead>")
        html.append("      <tr>")
        html.append("        <th>Nº</th>")
        html.append("        <th>Tribunal</th>")
        html.append("        <th>Data</th>")
        html.append("        <th>OAB</th>")
        html.append("        <th>Score</th>")
        html.append("        <th>Tipo</th>")
        html.append("        <th>Contexto</th>")
        html.append("      </tr>")
        html.append("    </thead>")
        html.append("    <tbody>")

        for i, match in enumerate(matches, 1):
            # Classe baseada em score
            score_class = self._get_score_class(match.score_relevancia)

            html.append("      <tr>")
            html.append(f"        <td>{i}</td>")
            html.append(f"        <td>{match.tribunal}</td>")
            html.append(f"        <td>{match.data_publicacao}</td>")
            html.append(f"        <td>{match.oab_numero}/{match.oab_uf}</td>")
            html.append(
                f"        <td class='{score_class}'>"
                f"{match.score_relevancia:.2f}</td>"
            )
            html.append(f"        <td>{match.tipo_ato or 'N/A'}</td>")
            html.append(f"        <td class='context'>{match.texto_contexto}</td>")
            html.append("      </tr>")

        html.append("    </tbody>")
        html.append("  </table>")

        html.append("</body>")
        html.append("</html>")

        # Escrever arquivo
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html))

        logger.info(f"HTML exportado: {output_path}")

    def _get_html_css(self) -> str:
        """Retorna CSS para HTML."""
        return """
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { color: #333; }
        .metadata { color: #666; font-size: 14px; }
        .stats { background: #f0f0f0; padding: 15px; margin: 20px 0; border-radius: 5px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th { background: #366092; color: white; padding: 10px; text-align: left; }
        td { padding: 8px; border-bottom: 1px solid #ddd; }
        tr:hover { background: #f5f5f5; }
        .score-high { background: #C6EFCE; font-weight: bold; }
        .score-medium { background: #FFEB9C; }
        .score-low { background: #FFC7CE; }
        .context { max-width: 400px; font-size: 12px; color: #555; }
        """

    def _get_score_class(self, score: float) -> str:
        """Retorna classe CSS baseada em score."""
        if score >= 0.8:
            return "score-high"
        elif score >= 0.6:
            return "score-medium"
        else:
            return "score-low"

    def _group_by_tribunal(self, matches: List) -> Dict[str, List]:
        """Agrupa matches por tribunal."""
        grouped = defaultdict(list)
        for match in matches:
            grouped[match.tribunal].append(match)
        return dict(grouped)

    def _group_by_oab(self, matches: List) -> Dict[Tuple[str, str], List]:
        """Agrupa matches por OAB."""
        grouped = defaultdict(list)
        for match in matches:
            key = (match.oab_numero, match.oab_uf)
            grouped[key].append(match)
        return dict(grouped)

    def _calculate_stats(self, matches: List) -> Dict:
        """Calcula estatísticas dos matches."""
        if not matches:
            return {
                'total_tribunais': 0,
                'total_oabs': 0,
                'score_medio': 0.0,
                'requer_revisao': 0,
                'por_tribunal': {}
            }

        tribunais = set(m.tribunal for m in matches)
        oabs = set((m.oab_numero, m.oab_uf) for m in matches)
        score_medio = sum(m.score_relevancia for m in matches) / len(matches)
        requer_revisao = sum(1 for m in matches if m.requer_revisao_manual)

        # Contar por tribunal
        por_tribunal = defaultdict(int)
        for match in matches:
            por_tribunal[match.tribunal] += 1

        return {
            'total_tribunais': len(tribunais),
            'total_oabs': len(oabs),
            'score_medio': score_medio,
            'requer_revisao': requer_revisao,
            'por_tribunal': dict(sorted(
                por_tribunal.items(),
                key=lambda x: x[1],
                reverse=True
            ))
        }


# ============================================================================
# EXEMPLO DE USO
# ============================================================================

if __name__ == "__main__":
    from dataclasses import dataclass
    from typing import List

    # Mock PublicacaoMatch para teste
    @dataclass
    class MockMatch:
        tribunal: str
        data_publicacao: str
        arquivo_pdf: str
        oab_numero: str
        oab_uf: str
        total_mencoes: int
        texto_contexto: str
        score_relevancia: float
        score_contexto: float = 0.5
        score_densidade: float = 0.5
        score_posicao: float = 0.5
        tipo_ato: Optional[str] = None
        requer_revisao_manual: bool = False

        def to_dict(self):
            return {
                'tribunal': self.tribunal,
                'data_publicacao': self.data_publicacao,
                'oab_numero': self.oab_numero,
                'oab_uf': self.oab_uf,
                'score_relevancia': self.score_relevancia
            }

    # Criar matches de exemplo
    matches = [
        MockMatch("TJSP", "2025-11-17", "/tmp/tjsp.pdf", "123456", "SP",
                  2, "Advogado Dr. João Silva...", 0.85, tipo_ato="Intimação"),
        MockMatch("STJ", "2025-11-17", "/tmp/stj.pdf", "789012", "RJ",
                  1, "Procurador Pedro Costa...", 0.65, tipo_ato="Sentença"),
        MockMatch("TRF1", "2025-11-16", "/tmp/trf1.pdf", "456789", "DF",
                  3, "Defensor Ana Paula...", 0.92, tipo_ato="Acórdão"),
    ]

    # Criar exporter
    exporter = ResultExporter(group_by_tribunal=True)

    # Exportar em todos os formatos
    output_dir = Path("/tmp/oab_exports")
    output_dir.mkdir(exist_ok=True)

    print("Exportando resultados...")

    exporter.export_json(matches, output_dir / "results.json")
    exporter.export_markdown(matches, output_dir / "results.md")
    exporter.export_txt(matches, output_dir / "results.txt")
    exporter.export_html(matches, output_dir / "results.html")

    try:
        exporter.export_excel(matches, output_dir / "results.xlsx")
    except ImportError:
        print("Skipping Excel export (openpyxl not installed)")

    print(f"\nResultados exportados em: {output_dir}")
